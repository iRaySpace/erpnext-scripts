import frappe
from openpyxl import load_workbook


def import_data():

    # Load the workbook
    wb = load_workbook("symptoms.xlsx")

    # Get the sheet
    sheet = wb.active

    # Max row
    max_row = sheet.max_row

    print "Importing the Data."

    # The index starts at 1
    for row in range(2, max_row + 1):

        # Symptom Name
        c1 = sheet.cell(row=row, column=1)

        # Description
        c2 = sheet.cell(row=row, column=2)

        # Causes
        c3 = sheet.cell(row=row, column=3)

        # When to See a Doctor
        c4 = sheet.cell(row=row, column=4)

        # Category
        c5 = sheet.cell(row=row, column=5)

        # Symptom's Description and Causes
        doc = frappe.get_doc({
            "doctype": "Symptoms Description and Causes",
            "symptom_name": c1.value,
            "description": c2.value,
            "causes": c3.value,
            "when": c4.value,
            "category": c5.value
        })

        doc.insert()

        print "Imported data {0}/{1}".format(row - 1, max_row - 1)
